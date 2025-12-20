from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime
from typing import Dict, List
import io
import platform
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, LineChart, Reference

# Register Japanese fonts
def register_japanese_fonts():
    """Register Japanese fonts for PDF generation"""
    try:
        system = platform.system()
        
        # Try to register common Japanese fonts
        font_paths = []
        
        if system == "Darwin":  # macOS
            # プロジェクトに含まれるフォントを優先的に使用
            import os
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
            project_font = os.path.join(base_dir, 'fonts', 'NotoSansCJK-Regular.ttf')
            
            font_paths = []
            # プロジェクトフォントを優先（存在する場合）
            if os.path.exists(project_font):
                font_paths.append(project_font)
            
            # ReportLabは.ttcファイルをサポートしていないため、.ttfファイルのみを使用
            font_paths.extend([
                "/System/Library/Fonts/Supplemental/AppleGothic.ttf",
                "/System/Library/Fonts/AppleGothic.ttf",
                "/System/Library/Fonts/Supplemental/NotoSansGothic-Regular.ttf",
            ])
        elif system == "Linux":
            font_paths = [
                "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
                "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
            ]
        elif system == "Windows":
            font_paths = [
                "C:/Windows/Fonts/msgothic.ttc",
                "C:/Windows/Fonts/msmincho.ttc",
            ]
        
        # Try to register the first available font
        for font_path in font_paths:
            try:
                pdfmetrics.registerFont(TTFont('Japanese', font_path))
                return 'Japanese'
            except (OSError, IOError, Exception):
                continue
        
        # If no system font found, try to use built-in font
        # ReportLab's built-in fonts don't support Japanese well
        # So we'll use a fallback approach
        return None
    except Exception:
        return None

# Register fonts on module import
JAPANESE_FONT = register_japanese_fonts()

class ReportService:
    @staticmethod
    def generate_pdf_report(
        user_name: str,
        analysis_data: Dict,
        summary_data: Dict,
        campaigns_data: List[Dict]
    ) -> bytes:
        """Generate PDF report"""
        
        buffer = io.BytesIO()
        # ページマージンを削減して2ページに収める
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch,
            leftMargin=0.6*inch,
            rightMargin=0.6*inch
        )
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        
        # Use Japanese font if available, otherwise use default
        font_name = JAPANESE_FONT if JAPANESE_FONT else 'Helvetica'
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=font_name,
            fontSize=26,
            textColor=colors.HexColor('#6366f1'),  # Indigo from dashboard
            spaceAfter=20,
            spaceBefore=0,
            alignment=TA_CENTER,
            leading=32
        )
        
        # heading_style は使用しない（spaceBefore=16ptが大きすぎる）
        # すべて normal_style ベースで作成
        # すべて normal_style ベースで作成
        
        # Create normal style with Japanese font
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=10,
            leading=14
        )
        
        # 統一されたセクション見出しスタイル（すべて normal_style ベース）
        section_heading_style = ParagraphStyle(
            'SectionHeading',
            parent=normal_style,
            fontSize=14,
            textColor=colors.HexColor('#111827'),
            spaceAfter=6,
            spaceBefore=0,
            leading=16
        )
        
        # Header with title and date (管理画面スタイル)
        from reportlab.platypus import KeepTogether
        # 分析レポートと生成日時を同じ行に表示
        header_text = f"分析レポート：生成日時：{datetime.now().strftime('%Y/%m/%d %H:%M:%S')}"
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=normal_style,
            fontName=font_name if font_name else 'Helvetica-Bold',
            fontSize=18,
            textColor=colors.HexColor('#111827'),
            spaceAfter=6,
            spaceBefore=0,
            leading=22
        )
        story.append(Paragraph(header_text, header_style))
        
        # 全体分析またはキャンペーン名を表示
        campaign_name = analysis_data.get('campaign_name')
        if campaign_name:
            campaign_text = f"キャンペーン: {campaign_name}"
        else:
            campaign_text = "全体分析"
        
        campaign_style = ParagraphStyle(
            'CampaignStyle',
            parent=normal_style,
            fontName=font_name,
            fontSize=12,
            textColor=colors.HexColor('#6366f1'),
            spaceAfter=8,
            spaceBefore=0,
            leading=16
        )
        story.append(Paragraph(campaign_text, campaign_style))
        story.append(Spacer(1, 0.15*inch))
        
        # AI Analysis Results (管理画面スタイル)
        if analysis_data.get('overall_rating'):
            # Overall Rating (管理画面スタイル)
            rating = analysis_data.get('overall_rating', 0)
            rating_stars = "★" * rating + "☆" * (5 - rating)
            
            rating_style = ParagraphStyle(
                'RatingStyle',
                parent=normal_style,
                fontSize=14,
                textColor=colors.HexColor('#111827'),
                spaceAfter=8,
                leading=18
            )
            
            story.append(Paragraph("<b>総合評価</b>", rating_style))
            story.append(Paragraph(f"{rating_stars}<br/>{rating}/5", rating_style))
            story.append(Spacer(1, 0.1*inch))
            
            comment = analysis_data.get('overall_comment', '')
            if comment:
                comment_style = ParagraphStyle(
                    'CommentStyle',
                    parent=normal_style,
                    fontSize=10,
                    textColor=colors.HexColor('#374151'),
                    leftIndent=0,
                    rightIndent=0,
                    spaceAfter=8,
                    leading=16
                )
                story.append(Paragraph(f'"{comment}"', comment_style))
                story.append(Spacer(1, 0.1*inch))
            
            # Issues (管理画面スタイル - ▲アイコン付き)
            issues = analysis_data.get('issues', [])
            if issues:
                story.append(Paragraph("▲ 主要課題", section_heading_style))
                
                for idx, issue in enumerate(issues):
                    # 最後の項目は spaceAfter=0
                    is_last = (idx == len(issues) - 1)
                    issue_style = ParagraphStyle(
                        'IssueStyle',
                        parent=normal_style,
                        fontSize=10,
                        textColor=colors.HexColor('#374151'),
                        leftIndent=0,
                        spaceAfter=0 if is_last else 6,
                        leading=16
                    )
                    
                    severity_symbol = {
                        '高': '▲',
                        '中': '◆',
                        '低': '◆'
                    }.get(issue.get('severity', ''), '◆')
                    
                    issue_text = f"""
                    <para>
                    <b>{severity_symbol} {issue.get('issue', '')}</b><br/>
                    {issue.get('impact', '')}
                    </para>
                    """
                    story.append(Paragraph(issue_text, issue_style))
                # 主要課題後のスペーサーなし
            
            # Recommendations (管理画面スタイル - カテゴリ別)
            recommendations = analysis_data.get('recommendations', [])
            if recommendations:
                story.append(Paragraph("◎ 改善提案詳細", section_heading_style))
                
                # カテゴリ別にグループ化
                by_category = {}
                for rec in recommendations:
                    category = rec.get('category', 'その他')
                    if category not in by_category:
                        by_category[category] = []
                    by_category[category].append(rec)
                
                category_keys = list(by_category.keys())
                for idx, (category, recs) in enumerate(by_category.items()):
                    is_last_category = (idx == len(category_keys) - 1)
                    category_style = ParagraphStyle(
                        'CategoryStyle',
                        parent=normal_style,
                        fontSize=11,
                        textColor=colors.HexColor('#6366f1'),
                        spaceAfter=3,
                        spaceBefore=0,
                        leading=16
                    )
                    story.append(Paragraph(f"<b>{category}</b>", category_style))
                    
                    for rec_idx, rec in enumerate(recs):
                        difficulty = rec.get('difficulty', 0)
                        difficulty_stars = "★" * difficulty + "☆" * (5 - difficulty)
                        
                        # 最後のカテゴリの最後の項目はスペースを0に
                        is_last_item = is_last_category and (rec_idx == len(recs) - 1)
                        rec_style = ParagraphStyle(
                            'RecStyle',
                            parent=normal_style,
                            fontSize=10,
                            textColor=colors.HexColor('#374151'),
                            leftIndent=0,
                            spaceAfter=0 if is_last_item else 4,
                            leading=16
                        )
                        
                        rec_text = f"""
                        <para>
                        <b>{rec.get('title', '')}</b><br/>
                        {rec.get('description', '')}<br/>
                        <br/>
                        <b>期待効果</b><br/>
                        {rec.get('expected_impact', '')}<br/>
                        <b>難易度</b> {difficulty_stars}
                        </para>
                        """
                        story.append(Paragraph(rec_text, rec_style))
                    # カテゴリ間のスペーサーを削除（すべて削除）
            
            # Action Plan (管理画面スタイル - ☐チェックボックス付き)
            action_plan = analysis_data.get('action_plan', [])
            if action_plan:
                story.append(Paragraph("☐ アクションプラン", section_heading_style))
                
                action_style = ParagraphStyle(
                    'ActionStyle',
                    parent=normal_style,
                    fontSize=10,
                    textColor=colors.HexColor('#374151'),
                    leftIndent=0,
                    spaceAfter=2,
                    leading=16
                )
                
                for i, action in enumerate(action_plan):
                    timeline = action.get('timeline', '')
                    responsible = action.get('responsible', '')
                    action_text = f"""
                    <para>
                    <b>{action.get('step', '')} {action.get('action', '')}</b><br/>
                    期間: {timeline} ▼担当: {responsible}
                    </para>
                    """
                    # 最後の項目はスペースを0に
                    if i == len(action_plan) - 1:
                        last_action_style = ParagraphStyle(
                            'LastActionStyle',
                            parent=action_style,
                            spaceAfter=0
                        )
                        story.append(Paragraph(action_text, last_action_style))
                    else:
                        story.append(Paragraph(action_text, action_style))
                # アクションプランとキャンペーン別パフォーマンスの間にスペーサーを追加しない
        
        # Campaign Performance Table (管理画面スタイル)
        if campaigns_data:
            story.append(Paragraph("📊 キャンペーン別パフォーマンス", section_heading_style))
            
            campaign_data = [['キャンペーン名', '費用', 'CV', 'ROAS', 'CPA']]
            for camp in campaigns_data[:10]:  # Top 10
                campaign_data.append([
                    str(camp.get('campaign_name', ''))[:30],
                    f"¥{camp.get('cost', 0):,.0f}",
                    str(camp.get('conversions', 0)),
                    f"{camp.get('roas', 0):.0f}%",
                    f"¥{camp.get('cpa', 0):,.0f}"
                ])
            
            campaign_table = Table(campaign_data, colWidths=[2.5*inch, 1*inch, 0.7*inch, 1*inch, 1*inch])
            campaign_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f9fafb')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#111827')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), font_name if font_name else 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), font_name if font_name else 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ]))
            
            story.append(campaign_table)
        
        # Build PDF
        doc.build(story)
        
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        return pdf_bytes
    
    @staticmethod
    def generate_excel_report(
        summary_data: Dict,
        campaigns_data: List[Dict],
        trends_data: List[Dict]
    ) -> bytes:
        """Generate Excel report"""
        
        output = io.BytesIO()
        wb = Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "サマリー"
        
        # Header
        ws_summary['A1'] = 'META広告分析レポート'
        ws_summary['A1'].font = Font(size=16, bold=True, color="1E40AF")
        ws_summary.merge_cells('A1:D1')
        
        ws_summary['A2'] = f"生成日時: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        period_start = summary_data.get('period', {}).get('start_date', '')
        period_end = summary_data.get('period', {}).get('end_date', '')
        ws_summary['A3'] = f"期間: {period_start} 〜 {period_end}"
        
        # KPI Table
        ws_summary['A5'] = '指標'
        ws_summary['B5'] = '値'
        ws_summary['A5'].font = Font(bold=True)
        ws_summary['B5'].font = Font(bold=True)
        
        totals = summary_data.get('totals', {})
        averages = summary_data.get('averages', {})
        
        kpi_rows = [
            ('総広告費', f"¥{totals.get('cost', 0):,.0f}"),
            ('総インプレッション', f"{totals.get('impressions', 0):,}"),
            ('総クリック数', f"{totals.get('clicks', 0):,}"),
            ('総コンバージョン数', f"{totals.get('conversions', 0):,}"),
            ('平均CTR', f"{averages.get('ctr', 0):.2f}%"),
            ('平均CPC', f"¥{averages.get('cpc', 0):,.0f}"),
            ('平均CPA', f"¥{averages.get('cpa', 0):,.0f}"),
            ('平均CVR', f"{averages.get('cvr', 0):.2f}%"),
            ('平均ROAS', f"{averages.get('roas', 0):.0f}%"),
        ]
        
        for i, (label, value) in enumerate(kpi_rows, start=6):
            ws_summary[f'A{i}'] = label
            ws_summary[f'B{i}'] = value
        
        # Campaign Data Sheet
        if campaigns_data:
            ws_campaigns = wb.create_sheet("キャンペーン詳細")
            
            headers = ['キャンペーン名', '費用', 'インプレッション', 'クリック', 'CV', 'CTR', 'CPC', 'CPA', 'CVR', 'ROAS']
            for col, header in enumerate(headers, start=1):
                cell = ws_campaigns.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            for row, camp in enumerate(campaigns_data, start=2):
                ws_campaigns.cell(row=row, column=1, value=camp.get('campaign_name', ''))
                ws_campaigns.cell(row=row, column=2, value=camp.get('cost', 0))
                ws_campaigns.cell(row=row, column=3, value=camp.get('impressions', 0))
                ws_campaigns.cell(row=row, column=4, value=camp.get('clicks', 0))
                ws_campaigns.cell(row=row, column=5, value=camp.get('conversions', 0))
                ws_campaigns.cell(row=row, column=6, value=camp.get('ctr', 0))
                ws_campaigns.cell(row=row, column=7, value=camp.get('cpc', 0))
                ws_campaigns.cell(row=row, column=8, value=camp.get('cpa', 0))
                ws_campaigns.cell(row=row, column=9, value=camp.get('cvr', 0))
                ws_campaigns.cell(row=row, column=10, value=camp.get('roas', 0))
            
            # Add chart if we have data
            if len(campaigns_data) > 0:
                try:
                    chart = BarChart()
                    chart.title = "キャンペーン別ROAS"
                    chart.x_axis.title = "キャンペーン"
                    chart.y_axis.title = "ROAS (%)"
                    
                    max_row = min(len(campaigns_data) + 1, 11)
                    data = Reference(ws_campaigns, min_col=10, min_row=1, max_row=max_row)
                    cats = Reference(ws_campaigns, min_col=1, min_row=2, max_row=max_row)
                    
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)
                    ws_campaigns.add_chart(chart, "L2")
                except Exception:
                    pass  # Skip chart if error
        
        # Trends Sheet (if data available)
        if trends_data and len(trends_data) > 0:
            ws_trends = wb.create_sheet("日別トレンド")
            
            trend_headers = ['日付', '費用', 'インプレッション', 'クリック', 'コンバージョン']
            for col, header in enumerate(trend_headers, start=1):
                cell = ws_trends.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            for row, trend in enumerate(trends_data, start=2):
                ws_trends.cell(row=row, column=1, value=trend.get('date', ''))
                ws_trends.cell(row=row, column=2, value=trend.get('cost', 0))
                ws_trends.cell(row=row, column=3, value=trend.get('impressions', 0))
                ws_trends.cell(row=row, column=4, value=trend.get('clicks', 0))
                ws_trends.cell(row=row, column=5, value=trend.get('conversions', 0))
            
            # Line chart for trends
            if len(trends_data) > 0:
                try:
                    line_chart = LineChart()
                    line_chart.title = "日別費用トレンド"
                    line_chart.x_axis.title = "日付"
                    line_chart.y_axis.title = "費用 (円)"
                    
                    data = Reference(ws_trends, min_col=2, min_row=1, max_row=len(trends_data)+1)
                    dates = Reference(ws_trends, min_col=1, min_row=2, max_row=len(trends_data)+1)
                    
                    line_chart.add_data(data, titles_from_data=True)
                    line_chart.set_categories(dates)
                    ws_trends.add_chart(line_chart, "G2")
                except Exception:
                    pass  # Skip chart if error
        
        # Auto-adjust column widths
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output)
        excel_bytes = output.getvalue()
        output.close()
        
        return excel_bytes







from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime
from typing import Dict, List
import io
import platform
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, LineChart, Reference

# Register Japanese fonts
def register_japanese_fonts():
    """Register Japanese fonts for PDF generation"""
    try:
        system = platform.system()
        
        # Try to register common Japanese fonts
        font_paths = []
        
        if system == "Darwin":  # macOS
            # プロジェクトに含まれるフォントを優先的に使用
            import os
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
            project_font = os.path.join(base_dir, 'fonts', 'NotoSansCJK-Regular.ttf')
            
            font_paths = []
            # プロジェクトフォントを優先（存在する場合）
            if os.path.exists(project_font):
                font_paths.append(project_font)
            
            # ReportLabは.ttcファイルをサポートしていないため、.ttfファイルのみを使用
            font_paths.extend([
                "/System/Library/Fonts/Supplemental/AppleGothic.ttf",
                "/System/Library/Fonts/AppleGothic.ttf",
                "/System/Library/Fonts/Supplemental/NotoSansGothic-Regular.ttf",
            ])
        elif system == "Linux":
            font_paths = [
                "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
                "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
            ]
        elif system == "Windows":
            font_paths = [
                "C:/Windows/Fonts/msgothic.ttc",
                "C:/Windows/Fonts/msmincho.ttc",
            ]
        
        # Try to register the first available font
        for font_path in font_paths:
            try:
                pdfmetrics.registerFont(TTFont('Japanese', font_path))
                return 'Japanese'
            except (OSError, IOError, Exception):
                continue
        
        # If no system font found, try to use built-in font
        # ReportLab's built-in fonts don't support Japanese well
        # So we'll use a fallback approach
        return None
    except Exception:
        return None

# Register fonts on module import
JAPANESE_FONT = register_japanese_fonts()

class ReportService:
    @staticmethod
    def generate_pdf_report(
        user_name: str,
        analysis_data: Dict,
        summary_data: Dict,
        campaigns_data: List[Dict]
    ) -> bytes:
        """Generate PDF report"""
        
        buffer = io.BytesIO()
        # ページマージンを削減して2ページに収める
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch,
            leftMargin=0.6*inch,
            rightMargin=0.6*inch
        )
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        
        # Use Japanese font if available, otherwise use default
        font_name = JAPANESE_FONT if JAPANESE_FONT else 'Helvetica'
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=font_name,
            fontSize=26,
            textColor=colors.HexColor('#6366f1'),  # Indigo from dashboard
            spaceAfter=20,
            spaceBefore=0,
            alignment=TA_CENTER,
            leading=32
        )
        
        # heading_style は使用しない（spaceBefore=16ptが大きすぎる）
        # すべて normal_style ベースで作成
        # すべて normal_style ベースで作成
        
        # Create normal style with Japanese font
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=10,
            leading=14
        )
        
        # 統一されたセクション見出しスタイル（すべて normal_style ベース）
        section_heading_style = ParagraphStyle(
            'SectionHeading',
            parent=normal_style,
            fontSize=14,
            textColor=colors.HexColor('#111827'),
            spaceAfter=6,
            spaceBefore=0,
            leading=16
        )
        
        # Header with title and date (管理画面スタイル)
        from reportlab.platypus import KeepTogether
        # 分析レポートと生成日時を同じ行に表示
        header_text = f"分析レポート：生成日時：{datetime.now().strftime('%Y/%m/%d %H:%M:%S')}"
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=normal_style,
            fontName=font_name if font_name else 'Helvetica-Bold',
            fontSize=18,
            textColor=colors.HexColor('#111827'),
            spaceAfter=6,
            spaceBefore=0,
            leading=22
        )
        story.append(Paragraph(header_text, header_style))
        
        # 全体分析またはキャンペーン名を表示
        campaign_name = analysis_data.get('campaign_name')
        if campaign_name:
            campaign_text = f"キャンペーン: {campaign_name}"
        else:
            campaign_text = "全体分析"
        
        campaign_style = ParagraphStyle(
            'CampaignStyle',
            parent=normal_style,
            fontName=font_name,
            fontSize=12,
            textColor=colors.HexColor('#6366f1'),
            spaceAfter=8,
            spaceBefore=0,
            leading=16
        )
        story.append(Paragraph(campaign_text, campaign_style))
        story.append(Spacer(1, 0.15*inch))
        
        # AI Analysis Results (管理画面スタイル)
        if analysis_data.get('overall_rating'):
            # Overall Rating (管理画面スタイル)
            rating = analysis_data.get('overall_rating', 0)
            rating_stars = "★" * rating + "☆" * (5 - rating)
            
            rating_style = ParagraphStyle(
                'RatingStyle',
                parent=normal_style,
                fontSize=14,
                textColor=colors.HexColor('#111827'),
                spaceAfter=8,
                leading=18
            )
            
            story.append(Paragraph("<b>総合評価</b>", rating_style))
            story.append(Paragraph(f"{rating_stars}<br/>{rating}/5", rating_style))
            story.append(Spacer(1, 0.1*inch))
            
            comment = analysis_data.get('overall_comment', '')
            if comment:
                comment_style = ParagraphStyle(
                    'CommentStyle',
                    parent=normal_style,
                    fontSize=10,
                    textColor=colors.HexColor('#374151'),
                    leftIndent=0,
                    rightIndent=0,
                    spaceAfter=8,
                    leading=16
                )
                story.append(Paragraph(f'"{comment}"', comment_style))
                story.append(Spacer(1, 0.1*inch))
            
            # Issues (管理画面スタイル - ▲アイコン付き)
            issues = analysis_data.get('issues', [])
            if issues:
                story.append(Paragraph("▲ 主要課題", section_heading_style))
                
                for idx, issue in enumerate(issues):
                    # 最後の項目は spaceAfter=0
                    is_last = (idx == len(issues) - 1)
                    issue_style = ParagraphStyle(
                        'IssueStyle',
                        parent=normal_style,
                        fontSize=10,
                        textColor=colors.HexColor('#374151'),
                        leftIndent=0,
                        spaceAfter=0 if is_last else 6,
                        leading=16
                    )
                    
                    severity_symbol = {
                        '高': '▲',
                        '中': '◆',
                        '低': '◆'
                    }.get(issue.get('severity', ''), '◆')
                    
                    issue_text = f"""
                    <para>
                    <b>{severity_symbol} {issue.get('issue', '')}</b><br/>
                    {issue.get('impact', '')}
                    </para>
                    """
                    story.append(Paragraph(issue_text, issue_style))
                # 主要課題後のスペーサーなし
            
            # Recommendations (管理画面スタイル - カテゴリ別)
            recommendations = analysis_data.get('recommendations', [])
            if recommendations:
                story.append(Paragraph("◎ 改善提案詳細", section_heading_style))
                
                # カテゴリ別にグループ化
                by_category = {}
                for rec in recommendations:
                    category = rec.get('category', 'その他')
                    if category not in by_category:
                        by_category[category] = []
                    by_category[category].append(rec)
                
                category_keys = list(by_category.keys())
                for idx, (category, recs) in enumerate(by_category.items()):
                    is_last_category = (idx == len(category_keys) - 1)
                    category_style = ParagraphStyle(
                        'CategoryStyle',
                        parent=normal_style,
                        fontSize=11,
                        textColor=colors.HexColor('#6366f1'),
                        spaceAfter=3,
                        spaceBefore=0,
                        leading=16
                    )
                    story.append(Paragraph(f"<b>{category}</b>", category_style))
                    
                    for rec_idx, rec in enumerate(recs):
                        difficulty = rec.get('difficulty', 0)
                        difficulty_stars = "★" * difficulty + "☆" * (5 - difficulty)
                        
                        # 最後のカテゴリの最後の項目はスペースを0に
                        is_last_item = is_last_category and (rec_idx == len(recs) - 1)
                        rec_style = ParagraphStyle(
                            'RecStyle',
                            parent=normal_style,
                            fontSize=10,
                            textColor=colors.HexColor('#374151'),
                            leftIndent=0,
                            spaceAfter=0 if is_last_item else 4,
                            leading=16
                        )
                        
                        rec_text = f"""
                        <para>
                        <b>{rec.get('title', '')}</b><br/>
                        {rec.get('description', '')}<br/>
                        <br/>
                        <b>期待効果</b><br/>
                        {rec.get('expected_impact', '')}<br/>
                        <b>難易度</b> {difficulty_stars}
                        </para>
                        """
                        story.append(Paragraph(rec_text, rec_style))
                    # カテゴリ間のスペーサーを削除（すべて削除）
            
            # Action Plan (管理画面スタイル - ☐チェックボックス付き)
            action_plan = analysis_data.get('action_plan', [])
            if action_plan:
                story.append(Paragraph("☐ アクションプラン", section_heading_style))
                
                action_style = ParagraphStyle(
                    'ActionStyle',
                    parent=normal_style,
                    fontSize=10,
                    textColor=colors.HexColor('#374151'),
                    leftIndent=0,
                    spaceAfter=2,
                    leading=16
                )
                
                for i, action in enumerate(action_plan):
                    timeline = action.get('timeline', '')
                    responsible = action.get('responsible', '')
                    action_text = f"""
                    <para>
                    <b>{action.get('step', '')} {action.get('action', '')}</b><br/>
                    期間: {timeline} ▼担当: {responsible}
                    </para>
                    """
                    # 最後の項目はスペースを0に
                    if i == len(action_plan) - 1:
                        last_action_style = ParagraphStyle(
                            'LastActionStyle',
                            parent=action_style,
                            spaceAfter=0
                        )
                        story.append(Paragraph(action_text, last_action_style))
                    else:
                        story.append(Paragraph(action_text, action_style))
                # アクションプランとキャンペーン別パフォーマンスの間にスペーサーを追加しない
        
        # Campaign Performance Table (管理画面スタイル)
        if campaigns_data:
            story.append(Paragraph("📊 キャンペーン別パフォーマンス", section_heading_style))
            
            campaign_data = [['キャンペーン名', '費用', 'CV', 'ROAS', 'CPA']]
            for camp in campaigns_data[:10]:  # Top 10
                campaign_data.append([
                    str(camp.get('campaign_name', ''))[:30],
                    f"¥{camp.get('cost', 0):,.0f}",
                    str(camp.get('conversions', 0)),
                    f"{camp.get('roas', 0):.0f}%",
                    f"¥{camp.get('cpa', 0):,.0f}"
                ])
            
            campaign_table = Table(campaign_data, colWidths=[2.5*inch, 1*inch, 0.7*inch, 1*inch, 1*inch])
            campaign_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f9fafb')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#111827')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), font_name if font_name else 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), font_name if font_name else 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ]))
            
            story.append(campaign_table)
        
        # Build PDF
        doc.build(story)
        
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        return pdf_bytes
    
    @staticmethod
    def generate_excel_report(
        summary_data: Dict,
        campaigns_data: List[Dict],
        trends_data: List[Dict]
    ) -> bytes:
        """Generate Excel report"""
        
        output = io.BytesIO()
        wb = Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "サマリー"
        
        # Header
        ws_summary['A1'] = 'META広告分析レポート'
        ws_summary['A1'].font = Font(size=16, bold=True, color="1E40AF")
        ws_summary.merge_cells('A1:D1')
        
        ws_summary['A2'] = f"生成日時: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        period_start = summary_data.get('period', {}).get('start_date', '')
        period_end = summary_data.get('period', {}).get('end_date', '')
        ws_summary['A3'] = f"期間: {period_start} 〜 {period_end}"
        
        # KPI Table
        ws_summary['A5'] = '指標'
        ws_summary['B5'] = '値'
        ws_summary['A5'].font = Font(bold=True)
        ws_summary['B5'].font = Font(bold=True)
        
        totals = summary_data.get('totals', {})
        averages = summary_data.get('averages', {})
        
        kpi_rows = [
            ('総広告費', f"¥{totals.get('cost', 0):,.0f}"),
            ('総インプレッション', f"{totals.get('impressions', 0):,}"),
            ('総クリック数', f"{totals.get('clicks', 0):,}"),
            ('総コンバージョン数', f"{totals.get('conversions', 0):,}"),
            ('平均CTR', f"{averages.get('ctr', 0):.2f}%"),
            ('平均CPC', f"¥{averages.get('cpc', 0):,.0f}"),
            ('平均CPA', f"¥{averages.get('cpa', 0):,.0f}"),
            ('平均CVR', f"{averages.get('cvr', 0):.2f}%"),
            ('平均ROAS', f"{averages.get('roas', 0):.0f}%"),
        ]
        
        for i, (label, value) in enumerate(kpi_rows, start=6):
            ws_summary[f'A{i}'] = label
            ws_summary[f'B{i}'] = value
        
        # Campaign Data Sheet
        if campaigns_data:
            ws_campaigns = wb.create_sheet("キャンペーン詳細")
            
            headers = ['キャンペーン名', '費用', 'インプレッション', 'クリック', 'CV', 'CTR', 'CPC', 'CPA', 'CVR', 'ROAS']
            for col, header in enumerate(headers, start=1):
                cell = ws_campaigns.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            for row, camp in enumerate(campaigns_data, start=2):
                ws_campaigns.cell(row=row, column=1, value=camp.get('campaign_name', ''))
                ws_campaigns.cell(row=row, column=2, value=camp.get('cost', 0))
                ws_campaigns.cell(row=row, column=3, value=camp.get('impressions', 0))
                ws_campaigns.cell(row=row, column=4, value=camp.get('clicks', 0))
                ws_campaigns.cell(row=row, column=5, value=camp.get('conversions', 0))
                ws_campaigns.cell(row=row, column=6, value=camp.get('ctr', 0))
                ws_campaigns.cell(row=row, column=7, value=camp.get('cpc', 0))
                ws_campaigns.cell(row=row, column=8, value=camp.get('cpa', 0))
                ws_campaigns.cell(row=row, column=9, value=camp.get('cvr', 0))
                ws_campaigns.cell(row=row, column=10, value=camp.get('roas', 0))
            
            # Add chart if we have data
            if len(campaigns_data) > 0:
                try:
                    chart = BarChart()
                    chart.title = "キャンペーン別ROAS"
                    chart.x_axis.title = "キャンペーン"
                    chart.y_axis.title = "ROAS (%)"
                    
                    max_row = min(len(campaigns_data) + 1, 11)
                    data = Reference(ws_campaigns, min_col=10, min_row=1, max_row=max_row)
                    cats = Reference(ws_campaigns, min_col=1, min_row=2, max_row=max_row)
                    
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)
                    ws_campaigns.add_chart(chart, "L2")
                except Exception:
                    pass  # Skip chart if error
        
        # Trends Sheet (if data available)
        if trends_data and len(trends_data) > 0:
            ws_trends = wb.create_sheet("日別トレンド")
            
            trend_headers = ['日付', '費用', 'インプレッション', 'クリック', 'コンバージョン']
            for col, header in enumerate(trend_headers, start=1):
                cell = ws_trends.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            for row, trend in enumerate(trends_data, start=2):
                ws_trends.cell(row=row, column=1, value=trend.get('date', ''))
                ws_trends.cell(row=row, column=2, value=trend.get('cost', 0))
                ws_trends.cell(row=row, column=3, value=trend.get('impressions', 0))
                ws_trends.cell(row=row, column=4, value=trend.get('clicks', 0))
                ws_trends.cell(row=row, column=5, value=trend.get('conversions', 0))
            
            # Line chart for trends
            if len(trends_data) > 0:
                try:
                    line_chart = LineChart()
                    line_chart.title = "日別費用トレンド"
                    line_chart.x_axis.title = "日付"
                    line_chart.y_axis.title = "費用 (円)"
                    
                    data = Reference(ws_trends, min_col=2, min_row=1, max_row=len(trends_data)+1)
                    dates = Reference(ws_trends, min_col=1, min_row=2, max_row=len(trends_data)+1)
                    
                    line_chart.add_data(data, titles_from_data=True)
                    line_chart.set_categories(dates)
                    ws_trends.add_chart(line_chart, "G2")
                except Exception:
                    pass  # Skip chart if error
        
        # Auto-adjust column widths
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output)
        excel_bytes = output.getvalue()
        output.close()
        
        return excel_bytes



